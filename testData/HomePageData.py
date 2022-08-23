import openpyxl

class HomePageData:

    test_HomePage_Data = [{"fname":"Akchhay","email":"Khanha", "pword":"123456", "Gender":"Male"}, {"fname":"Twinkle","email":"Khanha", "pword":"123456", "Gender":"Female"}]
    # in order to call method in the class, we need to create an object of that class and then call as object.method
    # in order to call method in the class directly as classname.method, we need to define method as static
    # Staticmethod inside the class does not need self parameter
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\prari\\Documents\\pythonExelDemo\\pythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]  # returns dictionary in the form of list which is the requirement in this case


