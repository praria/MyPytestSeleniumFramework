import openpyxl
book = openpyxl.load_workbook("C:\\Users\\prari\\Documents\\pythonExelDemo\\pythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = "Amit"
print(sheet.cell(row=2, column=5).value)

print(sheet['A3'].value)

Dict = {}


for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase2":

        for j in range(2, sheet.max_column+1):

            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)